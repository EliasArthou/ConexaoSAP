"""
Rotinas do SAP serão executadas por aqui
"""
import messagebox
import sys
import datetime
import time
from openpyxl import Workbook, load_workbook
import os
import auxiliares as aux
import pandas as pd


# from janela import App

def programarSQVI(transacoes, se, visual):
    """
    :param visual: tela de retorno do usuário.
    :param transacoes: lista contendo o dicionário com a transação e a classificação dos itens pesquisados.
    :param se: sessão do SAP que as ações do SAP será executado
    """

    try:
        # Variável para controlar se o Tipo mudou em relação à transação anterior para saber se será necessário um novo SELECT
        tipoatual = ''
        # Lista que receberá o resultado do SELECT
        lista = []
        # Inicia a variável
        datainicio = ''

        # Variável que receberá o intervalo de itens para a execução de cada JOB
        intervalo = aux.criarinputbox('Quantidade de Itens', 'Insira a quantidade de itens por job a ser extraído:',
                                      valorinicial='10000')
        # Verifica se foi digitado corretamente
        if len(intervalo) > 0:
            if intervalo.isnumeric():
                # Formata a entrada para número
                intervalo = int(intervalo)
        else:
            # Mensagem de erro e finalização do processo se o intervalo foi digitado errado
            messagebox.msgbox('Valor inválido para a quantidade de itens!', messagebox.MB_OK,
                              'Erro quantidade de itens')
            sys.exit()

        visual.acertaconfjanela(True)

        # ‘Looping’ para executar todas as visões solicitadas pelo usuário
        for index, transacao in enumerate(transacoes):
            # ==================== Parte Gráfica =======================================================
            # Coloca o nome do JOB na tela (da View do JOB, na verdade, visto que a transação é a SQVI)
            visual.mudartexto('labeljob', transacao[list(transacao)[0]])
            # Diz qual o número da transação (View) está sendo executada no momento no total de transações (Views)
            visual.mudartexto('statustrans', 'Item ' + str(index + 1) + ' de ' + str(len(transacoes)) + '...')
            # Atualiza a barra de progresso das transações (Views)
            visual.configurarbarra('barratrans', len(transacoes), index + 1)
            # ==================== Parte Gráfica =======================================================
            # Seleciona a transação desejada
            se.findById("wnd[0]/tbar[0]/okcd").text = "SQVI"
            # Confirma a transação
            se.findById("wnd[0]").sendVKey(0)
            # Seleciona a View
            se.findById("wnd[0]/usr/ctxtRS38R-QNUM").text = transacao[list(transacao)[0]]
            # Confirma a seleção
            se.findById("wnd[0]/usr/btnP1").press()

            # Conexão com o Banco de Dados
            conec = aux.Conec()

            if transacao[list(transacao)[1]] != tipoatual:
                # Pega o campo tipo pelo index dele (transforma os índices do dicionário numa lista e
                # depois pega o segundo índice da lista, ressaltando que numa lista o primeiro índice
                # é de valor 0)
                tipoatual = transacao[list(transacao)[1]]
                # Consulta a ser realizada no banco (será usado para fazer a seleção no SAP)
                visual.mudartexto('labelpassos', 'Executando Consulta no Banco...')
                lista = conec.consulta(
                    "SELECT DISTINCT [Nº doc.ref] FROM [BDSIRI].[UsrBDSIRI].[GIG Analise Compromisso]"
                    " WHERE UPPER ([Ctg.val])='" + tipoatual + "' ORDER BY [Nº doc.ref]")

            # Verifica se a lista veio com itens
            if len(lista) == 0:
                messagebox.msgbox('Sem item para analisar com o tipo informado!', messagebox.MB_OK, 'Tabela Vazia')
            else:
                # Retorna o que está fazendo em 'background'
                visual.mudartexto('labelpassos', 'Programando os JOBs...')
                # Quebra a lista em lista menores para o tamanho com a quantidade de item definido na variável intervalo
                sublistas = list(aux.chunks(lista, intervalo))
                indice = 0
                visual.mudartexto('labeljob', transacao[list(transacao)[0]])
                visual.configurarbarra('barrajob', len(sublistas), indice)

                # 'Looping' para quebrar o 'job' (view) em vários 'jobs'
                for indice, item in enumerate(sublistas):
                    visual.mudartexto('statusjobs', 'Item ' + str(indice + 1) + ' de ' + str(len(sublistas)) + '...')
                    # Atualiza a barra de progresso dos 'jobs' programados
                    visual.configurarbarra('barrajob', len(sublistas), indice)
                    # Grava a data e hora que começou a rodar os 'JOBs'
                    datainicio = datetime.datetime.now()
                    # Carrega os n itens (definido na variável intervalo) para a memória para ser "colado" depois
                    aux.list_to_clipboard(item, indice + 1)
                    if tipoatual == 'PEDIDOS':
                        # Para abrir a lista de colagem a caixa de texto não pode estar vazia
                        # (problema dessa transação específica)
                        se.findById("wnd[0]/usr/ctxtEBELN-LOW").text = "0"
                        # Abre a lista de colagem de itens da memória
                        se.findById("wnd[0]/usr/btn%_EBELN_%_APP_%-VALU_PUSH").press()
                    else:
                        # Para abrir a lista de colagem a caixa de texto não pode estar vazia
                        # (problema dessa transação específica)
                        se.findById("wnd[0]/usr/ctxtBANFN-LOW").text = "0"
                        # Abre a lista de colagem de itens da memória
                        se.findById("wnd[0]/usr/btn%_BANFN_%_APP_%-VALU_PUSH").press()

                    # Abre a janela para colar as informações da memória
                    if se.ActiveWindow.name == "wnd[1]":
                        # Limpa a lista
                        se.findById("wnd[1]/tbar[0]/btn[16]").press()
                        # Carrega a lista com os itens da memória
                        se.findById("wnd[1]/tbar[0]/btn[24]").press()
                        # "Confirma" a lista
                        se.findById("wnd[1]/tbar[0]/btn[8]").press()
                    # Selecionar o executar em 'background'
                    if se.ActiveWindow.name == "wnd[0]":
                        # Aperta o botão de programar 'JOBs'
                        se.findById("wnd[0]/mbar/menu[0]/menu[2]").select()
                        if se.ActiveWindow.name == "wnd[1]":
                            # Confirma a "inpressora" utilizada (normalmente só virtual)
                            se.findById("wnd[1]/tbar[0]/btn[13]").press()
                            # Seleciona que é imediata
                            se.findById("wnd[1]/usr/btnSOFORT_PUSH").press()
                            # Salva o 'JOB'
                            se.findById("wnd[1]/tbar[0]/btn[11]").press()
                # Retorna o que está fazendo em 'background'
                visual.mudartexto('labelpassos', 'Programação dos JOBs concluída!')
                # Sai da transação
                se.EndTransaction()
                # Grava a data e hora que terminou de rodar os 'JOBs'
                datafim = datetime.datetime.now()
                # Verifica se o arquivo de LOG existe
                if not os.path.isfile('JobLog.xlsx'):
                    # Cria o arquivo em memória
                    wb = Workbook()
                    # Pega a planilha aberta
                    ws = wb.active
                    # Adiciona o cabeçalho
                    ws.append(['Data Início', 'Hora Início', 'Data Fim', 'Hora Fim', 'Usuário', 'View'])
                    # Salva o arquivo
                    wb.save('JobLog.xlsx')

                # Verifica se o arquivo de LOG existe para não ter erro quando abrir o arquivo para salvar o LOG
                if os.path.isfile('JobLog.xlsx'):
                    # Carrega o arquivo na memória
                    wb = load_workbook('JobLog.xlsx')
                    # Pega a planilha aberta
                    ws = wb.active

                    # Salva as informações relevantes para resgastar os jobs no próximo passo
                    ws.append([datainicio.strftime("%d/%m/%Y"), datainicio.strftime("%X"), datafim.strftime("%d/%m/%Y"),
                               datafim.strftime("%X"), se.info.user, transacao[list(transacao)[0]]])
                    # Salva o arquivo com as alterações
                    wb.save('JobLog.xlsx')
    # Tratamento de Erros
    except Exception as e:
        messagebox.msgbox(str(e.message), messagebox.MB_OK, 'Erro')

    finally:
        # Trata a finalização do SAP
        if se is not None:
            se.finalizarsap()


def retornarjobs(transacoes, se, visual):
    """

    :param transacoes: lista de transações.
    :param se: sessão do SAP.
    :param visual: tela de “app” para ser atualizado.
    """
    # try:
    # Busca o arquivo de LOG (necessário para "guiar" a busca de JOBs na SM37)

    listacabecalho = []
    listavalor = []
    listachecks = []
    indicecabecalho = -1
    limitecharleft = 0
    indicecharleft = []
    achoucabecalho = False

    if os.path.isfile('JobLog.xlsx'):
        df = pd.read_excel('JobLog.xlsx')
    else:
        messagebox.msgbox('Arquivo de LOG não encontrado!', messagebox.MB_OK, 'Sem Arquivo de LOG')
        sys.exit()
    for indice, transacao in enumerate(transacoes):
        linha = len(transacoes) - indice
        # Seleciona a transação desejada
        se.findById("wnd[0]/tbar[0]/okcd").text = "SM37"
        # Confirma a transação
        se.findById("wnd[0]").sendVKey(0)
        se.findById("wnd[0]/usr/txtBTCH2170-JOBNAME").Text = "*" + transacao[list(transacao)[0]] + "*"
        se.findById("wnd[0]/usr/txtBTCH2170-USERNAME").Text = df['Usuário'].iloc[linha * -1]
        datainicio = df['Data Início'].iloc[linha * -1]
        se.findById("wnd[0]/usr/ctxtBTCH2170-FROM_DATE").Text = datainicio.replace('/', '.')
        # datafim = df['Data Fim'].iloc[-1]
        se.findById("wnd[0]/usr/ctxtBTCH2170-TO_DATE").Text = datainicio.replace('/', '.')
        horainicio = df['Hora Início'].iloc[linha * -1]
        se.findById("wnd[0]/usr/ctxtBTCH2170-FROM_TIME").Text = horainicio
        horafim = datetime.datetime(*time.strptime(horainicio, '%H:%M:%S')[:6]) + datetime.timedelta(minutes=5)
        horafim = horafim.time()
        se.findById("wnd[0]/usr/ctxtBTCH2170-TO_TIME").Text = str(horafim)
        se.findById("wnd[0]").sendVKey(8)
        textomensagem = se.findById("wnd[0]/sbar").Text
        textomensagem = str(textomensagem).strip()
        if len(textomensagem) == 0 or textomensagem != "Nenhum job corresponde às condições de seleção":
            telaSAP = se.findById("wnd[0]/usr")
            # Guarda a barra de rolagem
            rolagem = se.findById("wnd[0]/usr").verticalScrollbar
            print(rolagem.position, rolagem.maximum, rolagem.pagesize)
            while rolagem.position <= rolagem.maximum:
                for item in telaSAP.children:
                    if hasattr(item, 'tooltip'):
                        # Guarda o valor do popup que aparece quando se passa o mouse no item,
                        # está sendo usado porque os únicos que tem essa opção são os itens de
                        # cabeçalho
                        valor = item.tooltip
                        # Verifica se a informação está preenchida
                        if len(valor.strip()) > 0:
                            achoucabecalho = True
                            # Guarda o cabeçalho numa lista de cabeçalhos
                            listacabecalho.append(valor)
                            # Verifica se já achou a linha do cabeçalho
                            if indicecabecalho == -1:
                                # Armazena a linha do cabeçalho
                                indicecabecalho = item.chartop

                            # Verifica a posição do último item da linha possível
                            if limitecharleft < item.charleft:
                                # Armazena o maior índice de coluna da linha pra saber o limite da mesma
                                limitecharleft = item.charleft

                            # Guarda a posição de todas as colunas de informações
                            indicecharleft.append(item.charleft)

                        else:
                            if achoucabecalho:
                                if str(item.type).upper() == "GUICHECKBOX":
                                    # Guarda os checkboxes
                                    listachecks.append((item.id, rolagem.position))
                                    print(item.charleft, item.chartop, item.id, item.text, item.tooltip, item.id)
                                else:
                                    # Armazena a informação
                                    print(item.text)
                                    listavalor.append(item.text)

                    # Armazena a informação
                    # if item.charleft in indicecharleft:
                    #    print(item.charleft, item.chartop, item.id, item.text, item.tooltip, item.id)

    # Tratamento de Erros
    # except Exception as e:
    #    messagebox.msgbox(str(e.message), messagebox.MB_OK, 'Erro')

    # finally:
    # Trata a finalização do SAP
    #    if se is not None:
    #        se.finalizarsap()
